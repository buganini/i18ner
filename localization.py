#!/usr/bin/env python3

# python3 localization.py output/

import os
import sys
from openpyxl import load_workbook
import re
from xml.sax.saxutils import escape as xml_escape
import json
import requests
import unicodedata
import xml.etree.ElementTree as ET

header_row = 0
cursive_main_lang = False

def cursive(s):
	for p,q in zip("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ","𝒶𝒷𝒸𝒹𝑒𝒻𝑔𝒽𝒾𝒿𝓀𝓁𝓂𝓃𝑜𝓅𝓆𝓇𝓈𝓉𝓊𝓋𝓌𝓍𝓎𝓏𝒜𝐵𝒞𝒟𝐸𝐹𝒢𝐻𝐼𝒥𝒦𝐿𝑀𝒩𝒪𝒫𝒬𝑅𝒮𝒯𝒰𝒱𝒲𝒳𝒴𝒵"):
		s = s.replace(p, q)
	return s

def is_en(s):
	for c in s:
		if unicodedata.category(c) in ("Lo",):
			# print(c, unicodedata.category(c))
			return False
	return True

# https://github.com/translate/translate/blob/master/translate/storage/aresource.py#L219
WHITESPACE = ' \n\t'  # Whitespace that we collapse.
MULTIWHITESPACE = re.compile('[ \n\t]{2}(?!\\\\n)')
def android_escape(text, quote_wrapping_whitespaces=True):
	"""Escape all the characters which need to be escaped in an Android XML
	file.
	:param text: Text to escape
	:param quote_wrapping_whitespaces: If True, heading and trailing
		   whitespaces will be quoted placing the entire resulting text in
		   double quotes.
	"""
	if text is None:
		return
	if len(text) == 0:
		return ''
	text = text.replace('\\', '\\\\')
	# This will add non intrusive real newlines to
	# ones in translation improving readability of result
	text = text.replace('\n', '\\n')
	text = text.replace('\t', '\\t')
	text = text.replace('\'', '\\\'')
	text = text.replace('"', '\\"')

	# @ needs to be escaped at start
	if text.startswith('@'):
		text = '\\@' + text[1:]
	# Quote strings with more whitespace
	if ((quote_wrapping_whitespaces and (text[0] in WHITESPACE or text[-1] in WHITESPACE))
			or len(MULTIWHITESPACE.findall(text))) > 0:
		return '"%s"' % text
	return text

def aescape(s):
	s = xml_escape(s)
	s = android_escape(s)
	if s in ("@", "?"):
		s = "\\" + s
	return s

def iescape(s):
	return json.dumps(s, ensure_ascii=False)[1:-1]

def strip_note(s):
	return re.sub(r"\([^()]*\)", "", s).strip()

def yes_or_no(question):
	while "the answer is invalid":
		reply = str(input(question+' (y/n): ')).lower().strip()
		if not reply:
			continue
		if reply[0] == 'y':
			return True
		if reply[0] == 'n':
			return False

ref_key = "Ref Key"
android_key = "Android"
android_folder_key = "Android folder"
android_file_key = "Android file"
android_arg_key = "Android arg"
android_default_name = "strings"
ios_key = "iOS"
ios_file_key = "iOS file"
ios_arg_key = "iOS arg"
ios_default_name = "Localizable"
json_key = "JSON"
json_file_key = "JSON file"
json_default_name = "i18n"
jsons_key = "JSONS"
py_key = "Python"
py_file_key = "Python file"
py_default_name = "i18n"
xliff_key = "XLIFF"

ios_locale_map = {
	"tw":"zh-Hant",
	"cn":"zh-Hans",
	"jp":"ja",
	"kr":"ko",
	"cz":"cs",
	"se":"sv",
}
android_locale_map = {
	"tw":"zh-rTW",
	"cn":"zh-rCN",
	"jp":"ja",
	"kr":"ko",
	"cz":"cs",
	"se":"sv",
	"pt-BR":"pt-rBR",
}

ARGUMENT = r"\{\{(.*?)\}\}"
BACKREF = r"%(.*?)%"

class Null(str):
	def split(self, *args):
		return self

	def replace(self, *args):
		return ""

	def __str__(self):
		return ""

	def __bool__(self):
		return False

	def __repr__(self):
		return "<null>"

class Sheet():
	def __init__(self, i, name, sheet):
		self.number = i
		self.name = name
		self.sheet = sheet
		self.nrows = sheet.max_row - (header_row + 1)
		self.ncols = sheet.max_column
		self.cols = {}
		self.dat = {}
		while sheet.cell(header_row+1, self.ncols).value is None and self.ncols > 1:
			self.ncols -= 1
		for c in range(0, self.ncols):
			value = strip_note(sheet.cell(header_row+1, c+1).value or "")
			self.cols[value] = c

	def hasCol(self, c):
		return c in self.cols

	def get(self, r, c, default=Null()):
		try:
			return self.dat[r,c]
		except:
			if type(c) is str:
				if c in self.cols:
					v = (self.sheet.cell(r+(header_row + 1 + 1), self.cols[c] + 1).value or "").strip()
					if v == "":
						return default
					else:
						return v
				else:
					return default
			else:
				v = (self.sheet.cell(r+(header_row + 1 + 1), c + 1).value or "").strip()
				if v == "":
					return default
				else:
					return v

	def set(self, r, c, v):
		self.dat[r,c] = v

class Reader():
	def __init__(self, infile, including_sheets):
		self.xls = load_workbook(filename = infile, data_only=True)
		if not including_sheets:
			for sheet in [self.xls[n] for n in self.xls.sheetnames]:
				if yes_or_no("Include {}?".format(sheet.title)):
					including_sheets.append(sheet.title)

		self._sheets = []
		for i,sheet in enumerate([self.xls[n] for n in self.xls.sheetnames]):
			if not sheet.title in including_sheets:
				continue
			self._sheets.append(Sheet(i, sheet.title, sheet))

	def sheets(self):
		return self._sheets

def set_kv(data, path, value, outlog, ctx):
	cur = data
	for k in path[:-1]:
		if not k in cur:
			cur[k] = {}
		if isinstance(cur, dict):
			cur = cur[k]
		else:
			outlog.write("\x1b[1;33m[WARN] key conflict for {0} key {1} at sheet {2}\x1b[m\n".format(*ctx))
	if isinstance(cur, dict):
		if path[-1] in cur and isinstance(cur[path[-1]], dict):
			outlog.write("\x1b[1;33m[WARN] key prefix conflict for {0} key {1} at sheet {2}\x1b[m\n".format(*ctx))
		else:
			cur[path[-1]] = value
	else:
		outlog.write("\x1b[1;33m[WARN] key conflict for {0} key {1} at sheet {2}\x1b[m\n".format(*ctx))

def conv(input_path, output_dir, outlog, main_lang_key="en", lang_key = [], including_sheets = [], args={}):
	aF = {}
	iF = {}
	commonJData = []
	jData = {}
	jsData = {}
	pData = {}
	xlfData = {}
	aKeys = set()
	iKeys = set()
	jKeys = set()
	jsKeys = set()
	pKeys = set()
	xlfKeys = set()
	ios_use_base = args.get("ios_use_base", False)

	reader = Reader(input_path, including_sheets)

	sheets = []
	for sheet in reader.sheets():
		if not sheet.hasCol(main_lang_key):
			outlog.write("\x1b[1;31m[ERROR] Skipping sheet [{0}] {1} Main language key column not found\x1b[m\n".format(sheet.number, sheet.name))
			continue
		sheets.append(sheet)

	# build refs map
	ref_key_map = {}
	for sheet in sheets:
		for r in range(sheet.nrows):
			value = sheet.get(r, ref_key)
			if value:
				ref_key_map[value] = r

	# fill blank args
	for sheet in sheets:
		for r in range(sheet.nrows):
			aArg = [x.strip() for x in sheet.get(r, android_arg_key).split(",")]
			if aArg:
				sheet.set(r, android_arg_key, aArg)
			else:
				sheet.set(r, android_arg_key, [])
			iArg = [x.strip() for x in sheet.get(r, ios_arg_key).split(",")]
			if iArg:
				sheet.set(r, ios_arg_key, iArg)
			else:
				sheet.set(r, ios_arg_key, [])


	# tokenize args
	for sheet in sheets:
		for r in range(sheet.nrows):
			for lang in [main_lang_key] + lang_key:
				value = sheet.get(r, lang)
				tokens = re.split(ARGUMENT, value)
				sheet.set(r, lang, tokens)

	# args interpolation for refs
	for sheet in sheets:
		for r in range(sheet.nrows):
			for lang in [main_lang_key] + lang_key:
				tokens = sheet.get(r, lang)
				for i, token in list(reversed(list(enumerate(tokens))))[0::2]:
					va = re.split(BACKREF, token)
					nva = [va[0]]

					for j in range(1,len(va),2):
						if va[j]:
							if va[j] in ref_key_map:
								nva.append(va[j])
								nva.append(va[j+1])
							else:
								outlog.write("\x1b[1;31m[WARN] Back reference {0} not found in language {1} at sheet {2}\x1b[m\n".format(ref, lang, sheet.name))
								nva[-1] = nva[-1] + "%" + va[j] + "%" + va[j+1]
						else:
							nva[-1] = nva[-1] + "%" + va[j+1]

					va = nva

					for j,ref in list(reversed(list(enumerate(va))))[1::2]:
						va = va[:j] + [Null()] + sheet.get(ref_key_map[ref], lang) + [Null()] + va[j+1:]

						for key,arg_key in ((android_key, android_arg_key), (ios_key, ios_arg_key)):
							if not sheet.get(r, key):
								continue
							args = sheet.get(ref_key_map[ref], arg_key)
							if args:
								this_args = sheet.get(r, arg_key)
								this_args = this_args[:i] + args + this_args[i+1:]
								sheet.set(r, arg_key, this_args)
					tokens = tokens[:i] + va + tokens[i+1:]
				sheet.set(r, lang, tokens)

	if not sheets:
		outlog.write("\x1b[1;31m[ERROR] No Data\x1b[m\n")

	for sheet in sheets:
		for r in range(sheet.nrows):
			argIndex = {}
			tokens = sheet.get(r, main_lang_key)
			for key in tokens[1::2]:
				if not key:
					continue
				if key not in argIndex:
					argIndex[key] = len(argIndex)

			folder = sheet.get(r, android_folder_key).strip("/")

			if folder != "":
				folder += "/"

			aKey = sheet.get(r, android_key)
			aArg = sheet.get(r, android_arg_key)
			if aKey:
				kk = (folder, aKey)
				if kk in aKeys:
					outlog.write("\x1b[1;33m[WARN] Duplicated Android key: {0}\x1b[m\n".format(kk))
				else:
					aKeys.add(kk)

			iKey = sheet.get(r, ios_key)
			iArg = sheet.get(r, ios_arg_key)
			if iKey:
				kk = iKey
				if kk in iKeys:
					outlog.write("\x1b[1;33m[WARN] Duplicated iOS key: {0}\x1b[m\n".format(kk))
				else:
					iKeys.add(kk)

			jKey = sheet.get(r, json_key)
			if jKey:
				kk = jKey
				if kk in jKeys:
					outlog.write("\x1b[1;33m[WARN] Duplicated JSON key: {0}\x1b[m\n".format(kk))
				else:
					jKeys.add(kk)

			jsKey = sheet.get(r, jsons_key)
			if jsKey:
				kk = jsKey
				if kk in jsKeys:
					outlog.write("\x1b[1;33m[WARN] Duplicated JSONs key: {0}\x1b[m\n".format(kk))
				else:
					jsKeys.add(kk)

			pKey = sheet.get(r, py_key)
			if pKey:
				kk = pKey
				if kk in pKeys:
					outlog.write("\x1b[1;33m[WARN] Duplicated Python key: {0}\x1b[m\n".format(kk))
				else:
					pKeys.add(kk)

			xlfKey = sheet.get(r, xliff_key)
			if xlfKey:
				kk = xlfKey
				if kk in xlfKeys:
					outlog.write("\x1b[1;33m[WARN] Duplicated XLIFF key: {0}\x1b[m\n".format(kk))
				else:
					xlfKeys.add(kk)

			for lang in [main_lang_key] + lang_key:
				value = sheet.get(r, lang)
				if not value:
					continue
				if len(value)==1 and not value[0]:
					continue

				if lang == main_lang_key and cursive_main_lang:
					for i in range(0, len(value), 2):
						value[i] = cursive(value[i])

				if aKey:
					# translate formatter
					va = list(value)
					for i in range(1, len(va), 2):
						if not va[i]:
							continue
						if va[i] in argIndex:
							ai = argIndex[va[i]]
							if ai < len(aArg):
								arg = aArg[ai]
							else:
								outlog.write("\x1b[1;31m[ERROR] Sheet \"{0}\": Undefined arg for Android key: {1}[{2}]\x1b[m\n".format(sheet.name, aKey, va[i]))
								return
							va[i] = "%{0}${1}".format(ai+1, arg)
						else:
							outlog.write("\x1b[1;31m[ERROR] Unexpected variable {0} for Android key {1} in language {2} at sheet {3}\x1b[m\n".format(va[i], aKey, lang, sheet.name))

					# escape data
					if not aArg == ["-"]:
						for i in range(0, len(va), 2):
							va[i] = va[i].replace("%", "%%")

					file = sheet.get(r, android_file_key, android_default_name)

					if lang == main_lang_key:
						aLang = ""
					else:
						aLang = "-" + android_locale_map.get(lang, lang)

					fk = (folder, aLang, file)
					if fk not in aF:
						aPath = os.path.join(output_dir, "android-strings/{0}values{1}/{2}.xml".format(folder, aLang, file))
						d = os.path.dirname(aPath)
						if not os.path.exists(d):
							os.makedirs(d)
						aF[fk] = open(aPath, "w", encoding="utf-8")
						aF[fk].write("<?xml version=\"1.0\" encoding=\"utf-8\"?>\n<resources>\n");

					s = "".join(va)
					if not s:
						continue

					if lang == "en" and not is_en(s):
						outlog.write("\x1b[1;33m[WARN] Non-English in EN string: Android/{0}: {1}\x1b[m\n".format(aKey, s))

					aF[fk].write("    <string name=\"{0}\">{1}</string>\n".format(aKey, aescape(s)))

				if iKey:
					# translate formatter
					va = list(value)
					for i in range(1, len(va), 2):
						if not va[i]:
							continue
						if va[i] in argIndex:
							ai = argIndex[va[i]]
							if ai < len(iArg):
								arg = iArg[ai]
							else:
								outlog.write("\x1b[1;31m[ERROR] Sheet \"{0}\": Undefined arg for iOS key: {1}[{2}]\x1b[m\n".format(sheet.name, iKey, va[i]))
								return
							va[i] = "%{0}${1}".format(ai+1, arg)
						else:
							outlog.write("\x1b[1;31m[ERROR] Unexpected variable {0} for iOS key {1} in language {2} at sheet {3}\x1b[m\n".format(va[i], iKey, lang, sheet.name))

					# escape data
					if not iArg == ["-"]:
						for i in range(0, len(va), 2):
							va[i] = va[i].replace("%", "%%")

					s = "".join(va)
					if not s:
						continue

					file = sheet.get(r, ios_file_key, ios_default_name)

					iLang = ios_locale_map.get(lang, lang)
					fk = (iLang, file)
					if fk not in iF:
						iPath = os.path.join(output_dir, "ios-strings/{0}.lproj/{1}.strings".format(iLang, file))
						d = os.path.dirname(iPath)
						if not os.path.exists(d):
							os.makedirs(d)
						iF[fk] = open(iPath, "w", encoding="utf-8")

					if lang == "en" and not is_en(s):
						outlog.write("\x1b[1;33m[WARN] Non-English in EN string: iOS/{0}: {1}\x1b[m\n".format(iKey, s))

					iF[fk].write("\"{0}\" = \"{1}\";\n".format(iKey, iescape(s)))

					if lang==main_lang_key and ios_use_base:
						iLang = "Base"
						base_fk = (iLang, file)
						if base_fk not in iF:
							iPath = os.path.join(output_dir, "ios-strings/{0}.lproj/{1}.strings".format(iLang, file))
							d = os.path.dirname(iPath)
							if not os.path.exists(d):
								os.makedirs(d)
							iF[base_fk] = open(iPath, "w", encoding="utf-8")

						iF[base_fk].write("\"{0}\" = \"{1}\";\n".format(iKey, iescape(s)))


				if jKey:
					va = list(value)
					for i in range(1, len(va), 2):
						if not va[i]:
							continue
						va[i] = "{{{}}}".format(va[i])

					s = "".join(va)
					if lang == "en" and not is_en(s):
						outlog.write("\x1b[1;33m[WARN] Non-English in EN string: JSON/{0}: {1}\x1b[m\n".format(jKey, s))
					jpath = [lang] + jKey.split(".")
					file = sheet.get(r, json_file_key, json_default_name)
					if file=="*":
						commonJData.append((jpath, s, ("JSON", jKey, sheet.name)))
					else:
						if not file in jData:
							jData[file] = []
						jData[file].append((jpath, s, ("JSON", jKey, sheet.name)))

				if jsKey:
					va = list(value)
					for i in range(1, len(va), 2):
						if not va[i]:
							continue
						va[i] = "{{{}}}".format(va[i])

					s = "".join(va)
					if lang == "en" and not is_en(s):
						outlog.write("\x1b[1;33m[WARN] Non-English in EN string: JSONs/{0}: {1}\x1b[m\n".format(jKey, s))
					jpath = jsKey.split(".")
					file = lang
					if not file in jsData:
						jsData[file] = {}
					set_kv(jsData[file], jpath, s, outlog, ("JSONs", jKey, sheet.name))

				if pKey:
					va = list(value)
					for i in range(1, len(va), 2):
						if not va[i]:
							continue
						va[i] = "{{{}}}".format(va[i])

					s = "".join(va)
					if lang == "en" and not is_en(s):
						outlog.write("\x1b[1;33m[WARN] Non-English in EN string: Python/{0}: {1}\x1b[m\n".format(pKey, s))
					ppath = [lang, pKey]
					file = sheet.get(r, py_file_key, py_default_name)
					if not file in pData:
						pData[file] = {}
					cur = pData[file]
					for k in ppath[:-1]:
						if not k in cur:
							cur[k] = {}
						if type(cur) is dict:
							cur = cur[k]
						else:
							outlog.write("\x1b[1;31m[ERROR] key conflict for Python key {0} at sheet {1}\x1b[m\n".format(pKey, sheet.name))
					if type(cur) is dict:
						cur[ppath[-1]] = s
					else:
						outlog.write("\x1b[1;31m[ERROR] key conflict for Python key {0} at sheet {1}\x1b[m\n".format(pKey, sheet.name))

				if xlfKey:
					va = list(value)
					for i in range(1, len(va), 2):
						if not va[i]:
							continue
						va[i] = "{{{{{}}}}}".format(va[i])

					s = "".join(va)
					if lang == "en" and not is_en(s):
						outlog.write("\x1b[1;33m[WARN] Non-English in EN string: XLIFF/{0}: {1}\x1b[m\n".format(xlfKey, s))
					file = lang
					if not file in xlfData:
						xlfData[file] = {}
					xlfData[file][xlfKey] = va

		print("Processed", sheet.name)

	for fk in aF:
		aF[fk].write("</resources>\n");
		aF[fk].close()

	for fk in iF:
		iF[fk].close()

	for fn in jData:
		data = {}
		for path, value, ctx in commonJData:
			set_kv(data, path, value, outlog, ctx)
		for path, value, ctx in jData[fn]:
			set_kv(data, path, value, outlog, ctx)

		jPath = os.path.join(output_dir, "json/{}.json".format(fn))
		d = os.path.dirname(jPath)
		if not os.path.exists(d):
			os.makedirs(d)
		with open(jPath, "w", encoding="utf-8") as f:
			json.dump(data, f, ensure_ascii=False, indent=4)

	for fn in jsData:
		jPath = os.path.join(output_dir, "jsons/{}.json".format(fn))
		d = os.path.dirname(jPath)
		if not os.path.exists(d):
			os.makedirs(d)
		with open(jPath, "w", encoding="utf-8") as f:
			json.dump(jsData[fn], f, ensure_ascii=False, indent=4)

	for fn in pData:
		pPath = os.path.join(output_dir, "{}.py".format(fn))
		d = os.path.dirname(pPath)
		if not os.path.exists(d):
			os.makedirs(d)
		with open(pPath, "w", encoding="utf-8") as f:
			f.write("I18N = ")
			f.write(repr(pData[fn]))
			f.write("\n")

	for fn in xlfData:
		xliff = ET.Element('xliff')
		xliff.set("version", "1.2")
		xliff.set("xmlns", "urn:oasis:names:tc:xliff:document:1.2")
		xliff.text = "\n"
		xliff.tail = "\n"
		file = ET.SubElement(xliff, 'file')
		file.set("source-language", "raw")
		file.set("datatype", "plaintext")
		file.set("original", "ng2.template")
		file.set("target-language", fn)
		file.text = "\n"
		file.tail = "\n"
		body = ET.SubElement(file, 'body')
		body.text = "\n\n"
		body.tail = "\n"
		for k in xlfData[fn]:
			tu = ET.SubElement(body, "trans-unit")
			tu.set("id", k)
			tu.set("datatype", "html")
			tu.text = "\n  "
			tu.tail = "\n\n"
			source = ET.SubElement(tu, "source")
			source.text = xlfData[main_lang_key][k][0]
			for i in range(1, len(xlfData[main_lang_key][k]), 2):
				x = ET.SubElement(source, "x")
				x.set("id", "INTERPOLATION")
				x.set("equiv-text", xlfData[main_lang_key][k][i])
				if i+1 < len(xlfData[main_lang_key][k]):
					x.tail = xlfData[main_lang_key][k][i+1]
			source.tail = "\n  "
			target = ET.SubElement(tu, "target")
			target.text = xlfData[fn][k][0]
			for i in range(1, len(xlfData[fn][k]), 2):
				x = ET.SubElement(target, "x")
				x.set("id", "INTERPOLATION")
				x.set("equiv-text", xlfData[fn][k][i])
				if i+1 < len(xlfData[fn][k]):
					x.tail = xlfData[fn][k][i+1]
			target.tail = "\n"
		outfile = f"xliff/messages.{fn}.xlf"
		d = os.path.dirname(outfile)
		if not os.path.exists(d):
			os.makedirs(d)
		with open(outfile, "wb") as f:
			f.write(ET.tostring(xliff, encoding="utf-8", xml_declaration=True))

if __name__ == "__main__":
	main_lang_key = "en"
	lang_key = [
		"tw",
		"ja",
		"ko",
		"de",
		"fr",
		"nl",
		"es",
		"ru",
		"th",
		"vi",
	]
	including_sheets = []

	outdir = sys.argv[1]
	if not os.path.exists(outdir):
		os.makedirs(outdir)

	gdoc = "https://docs.google.com/spreadsheets/d/1FDASCVNyga8KtwzxSj0PHwr7KyRrUKouwgexpfVYn2s/export?format=xlsx"
	req = requests.get(gdoc)
	f = os.path.join(outdir, "localization.xlsx")
	fp = open(f, "wb")
	fp.write(req.content)
	fp.close()
	conv(f, outdir, sys.stdout, main_lang_key, lang_key, including_sheets)
